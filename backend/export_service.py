"""
Export Service für PDF und Excel Reports
"""

import io
from datetime import datetime
from typing import Dict, List
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class PDFExportService:
    """Service für PDF-Export von Lohnsteuerberechnungen"""

    def __init__(self):
        self.styles = getSampleStyleSheet()
        self.title_style = ParagraphStyle(
            "CustomTitle",
            parent=self.styles["Heading1"],
            fontSize=18,
            spaceAfter=30,
            textColor=colors.darkblue,
        )

    def generate_calculation_report(
        self, input_data: Dict, result: Dict, calculation_id: str = None
    ) -> bytes:
        """Generiert einen PDF-Report für eine Lohnsteuerberechnung"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=18,
        )

        # Story für PDF-Inhalt
        story = []

        # Titel
        title = Paragraph("Lohnsteuerberechnung 2025", self.title_style)
        story.append(title)
        story.append(Spacer(1, 12))

        # Berechnungsdatum und ID
        date_str = datetime.now().strftime("%d.%m.%Y %H:%M")
        if calculation_id:
            info_text = (
                f"Berechnung vom: {date_str}<br/>Berechnungs-ID: {calculation_id}"
            )
        else:
            info_text = f"Berechnung vom: {date_str}"

        info_para = Paragraph(info_text, self.styles["Normal"])
        story.append(info_para)
        story.append(Spacer(1, 20))

        # Eingabedaten Tabelle
        story.append(Paragraph("Eingabedaten", self.styles["Heading2"]))
        input_table_data = self._prepare_input_table_data(input_data)
        input_table = Table(input_table_data, colWidths=[3 * inch, 2 * inch])
        input_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 12),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )
        story.append(input_table)
        story.append(Spacer(1, 20))

        # Ergebnisse Tabelle
        story.append(Paragraph("Berechnungsergebnis", self.styles["Heading2"]))
        result_table_data = self._prepare_result_table_data(input_data, result)
        result_table = Table(result_table_data, colWidths=[3 * inch, 2 * inch])
        result_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 12),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.lightblue),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                    (
                        "FONTNAME",
                        (0, -1),
                        (-1, -1),
                        "Helvetica-Bold",
                    ),  # Letzte Zeile fett
                    (
                        "BACKGROUND",
                        (0, -1),
                        (-1, -1),
                        colors.yellow,
                    ),  # Letzte Zeile hervorheben
                ]
            )
        )
        story.append(result_table)
        story.append(Spacer(1, 20))

        # Jahreshochrechnung
        if input_data.get("LZZ") != 1:  # Nicht bei jährlicher Berechnung
            story.append(Paragraph("Jahreshochrechnung", self.styles["Heading2"]))
            yearly_table_data = self._prepare_yearly_projection(input_data, result)
            yearly_table = Table(yearly_table_data, colWidths=[3 * inch, 2 * inch])
            yearly_table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.darkgreen),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, 0), 12),
                        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                        ("BACKGROUND", (0, 1), (-1, -1), colors.lightgreen),
                        ("GRID", (0, 0), (-1, -1), 1, colors.black),
                    ]
                )
            )
            story.append(yearly_table)
            story.append(Spacer(1, 20))

        # Hinweise
        story.append(Paragraph("Wichtige Hinweise", self.styles["Heading2"]))
        hinweise = [
            "• Diese Berechnung erfolgt nach dem Programmablaufplan (PAP) 2025",
            "• Alle Angaben ohne Gewähr",
            "• Sozialversicherungsbeiträge sind nicht enthalten",
            "• Bei Fragen wenden Sie sich an Ihren Steuerberater",
        ]
        for hinweis in hinweise:
            story.append(Paragraph(hinweis, self.styles["Normal"]))

        # PDF generieren
        doc.build(story)
        pdf_data = buffer.getvalue()
        buffer.close()

        return pdf_data

    def _prepare_input_table_data(self, input_data: Dict) -> List[List[str]]:
        """Bereitet Eingabedaten für Tabelle vor"""
        data = [["Parameter", "Wert"]]

        # Wichtigste Parameter anzeigen
        params = {
            "RE4": f"{float(input_data.get('RE4', 0)) / 100:.2f} €",
            "STKL": self._get_steuerklasse_text(input_data.get("STKL")),
            "LZZ": self._get_lzz_text(input_data.get("LZZ")),
            "ZKF": str(input_data.get("ZKF", 0)),
            "R": self._get_religion_text(input_data.get("R", 0)),
            "KVZ": f"{float(input_data.get('KVZ', 0)):.1f} %",
            "PKV": self._get_pkv_text(input_data.get("PKV", 0)),
        }

        param_names = {
            "RE4": "Bruttolohn",
            "STKL": "Steuerklasse",
            "LZZ": "Zeitraum",
            "ZKF": "Kinderfreibeträge",
            "R": "Kirchensteuer",
            "KVZ": "KV-Zusatzbeitrag",
            "PKV": "Krankenversicherung",
        }

        for key, value in params.items():
            data.append([param_names[key], value])

        return data

    def _prepare_result_table_data(
        self, input_data: Dict, result: Dict
    ) -> List[List[str]]:
        """Bereitet Ergebnisdaten für Tabelle vor"""
        data = [["Abzug", "Betrag"]]

        lstlzz = float(result.get("LSTLZZ", 0)) / 100
        solzlzz = float(result.get("SOLZLZZ", 0)) / 100
        bk = float(result.get("BK", 0)) / 100
        brutto = float(input_data.get("RE4", 0)) / 100

        total_tax = lstlzz + solzlzz + bk
        netto = brutto - total_tax

        data.extend(
            [
                ["Lohnsteuer", f"{lstlzz:.2f} €"],
                ["Solidaritätszuschlag", f"{solzlzz:.2f} €"],
                ["Kirchensteuer", f"{bk:.2f} €"],
                ["Gesamte Steuerbelastung", f"{total_tax:.2f} €"],
                ["Nettolohn", f"{netto:.2f} €"],
            ]
        )

        return data

    def _prepare_yearly_projection(
        self, input_data: Dict, result: Dict
    ) -> List[List[str]]:
        """Bereitet Jahreshochrechnung vor"""
        data = [["Jahreswerte", "Betrag"]]

        multiplier = self._get_yearly_multiplier(input_data.get("LZZ"))

        lstlzz = float(result.get("LSTLZZ", 0)) / 100 * multiplier
        solzlzz = float(result.get("SOLZLZZ", 0)) / 100 * multiplier
        bk = float(result.get("BK", 0)) / 100 * multiplier
        brutto = float(input_data.get("RE4", 0)) / 100 * multiplier

        total_tax = lstlzz + solzlzz + bk
        netto = brutto - total_tax
        effective_rate = (total_tax / brutto * 100) if brutto > 0 else 0

        data.extend(
            [
                ["Jahresbruttolohn", f"{brutto:.2f} €"],
                ["Jährliche Steuerbelastung", f"{total_tax:.2f} €"],
                ["Jährlicher Nettolohn", f"{netto:.2f} €"],
                ["Effektiver Steuersatz", f"{effective_rate:.1f} %"],
            ]
        )

        return data

    def _get_yearly_multiplier(self, lzz: int) -> int:
        """Gibt Multiplikator für Jahreshochrechnung zurück"""
        multipliers = {1: 1, 2: 12, 3: 52, 4: 365}
        return multipliers.get(lzz, 12)

    def _get_steuerklasse_text(self, stkl: int) -> str:
        """Konvertiert Steuerklasse zu Text"""
        classes = {
            1: "I (ledig)",
            2: "II (alleinerziehend)",
            3: "III (verheiratet, höheres Einkommen)",
            4: "IV (verheiratet, ähnliches Einkommen)",
            5: "V (verheiratet, geringeres Einkommen)",
            6: "VI (Nebenjob)",
        }
        return classes.get(stkl, str(stkl))

    def _get_lzz_text(self, lzz: int) -> str:
        """Konvertiert LZZ zu Text"""
        periods = {1: "Jahr", 2: "Monat", 3: "Woche", 4: "Tag"}
        return periods.get(lzz, str(lzz))

    def _get_religion_text(self, r: int) -> str:
        """Konvertiert Religionsgemeinschaft zu Text"""
        religions = {0: "Keine", 1: "Evangelisch/Katholisch (9%)", 2: "Andere (8%)"}
        return religions.get(r, str(r))

    def _get_pkv_text(self, pkv: int) -> str:
        """Konvertiert PKV zu Text"""
        types = {
            0: "Gesetzlich",
            1: "Privat ohne AG-Zuschuss",
            2: "Privat mit AG-Zuschuss",
        }
        return types.get(pkv, str(pkv))


class ExcelExportService:
    """Service für Excel-Export von Lohnsteuerberechnungen"""

    def generate_calculation_report(
        self, input_data: Dict, result: Dict, calculation_id: str = None
    ) -> bytes:
        """Generiert einen Excel-Report für eine Lohnsteuerberechnung"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Lohnsteuerberechnung"

        # Styles definieren
        title_font = Font(size=16, bold=True, color="1F4E79")
        header_font = Font(size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        data_font = Font(size=10)
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Titel
        ws["A1"] = "Lohnsteuerberechnung 2025"
        ws["A1"].font = title_font
        ws.merge_cells("A1:D1")

        # Datum und ID
        row = 3
        ws[f"A{row}"] = f"Berechnung vom: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        if calculation_id:
            row += 1
            ws[f"A{row}"] = f"Berechnungs-ID: {calculation_id}"

        # Eingabedaten
        row += 3
        ws[f"A{row}"] = "Eingabedaten"
        ws[f"A{row}"].font = header_font
        ws[f"A{row}"].fill = header_fill
        ws[f"B{row}"].fill = header_fill
        ws.merge_cells(f"A{row}:B{row}")

        row += 1
        input_data_rows = self._prepare_excel_input_data(input_data)
        for param, value in input_data_rows:
            ws[f"A{row}"] = param
            ws[f"B{row}"] = value
            ws[f"A{row}"].font = data_font
            ws[f"B{row}"].font = data_font
            ws[f"A{row}"].border = border
            ws[f"B{row}"].border = border
            row += 1

        # Ergebnisse
        row += 2
        ws[f"A{row}"] = "Berechnungsergebnis"
        ws[f"A{row}"].font = header_font
        ws[f"A{row}"].fill = header_fill
        ws[f"B{row}"].fill = header_fill
        ws.merge_cells(f"A{row}:B{row}")

        row += 1
        result_data_rows = self._prepare_excel_result_data(input_data, result)
        for param, value in result_data_rows:
            ws[f"A{row}"] = param
            ws[f"B{row}"] = value
            ws[f"A{row}"].font = data_font
            ws[f"B{row}"].font = data_font
            ws[f"A{row}"].border = border
            ws[f"B{row}"].border = border
            if "Nettolohn" in param or "Gesamte" in param:
                ws[f"A{row}"].font = Font(size=10, bold=True)
                ws[f"B{row}"].font = Font(size=10, bold=True)
            row += 1

        # Spaltenbreite anpassen
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20

        # Excel-Datei in Bytes konvertieren
        buffer = io.BytesIO()
        wb.save(buffer)
        excel_data = buffer.getvalue()
        buffer.close()

        return excel_data

    def _prepare_excel_input_data(self, input_data: Dict) -> List[tuple]:
        """Bereitet Eingabedaten für Excel vor"""
        pdf_service = PDFExportService()

        return [
            ("Bruttolohn", f"{float(input_data.get('RE4', 0)) / 100:.2f} €"),
            (
                "Steuerklasse",
                pdf_service._get_steuerklasse_text(input_data.get("STKL")),
            ),
            ("Zeitraum", pdf_service._get_lzz_text(input_data.get("LZZ"))),
            ("Kinderfreibeträge", str(input_data.get("ZKF", 0))),
            ("Kirchensteuer", pdf_service._get_religion_text(input_data.get("R", 0))),
            ("KV-Zusatzbeitrag", f"{float(input_data.get('KVZ', 0)):.1f} %"),
            (
                "Krankenversicherung",
                pdf_service._get_pkv_text(input_data.get("PKV", 0)),
            ),
        ]

    def _prepare_excel_result_data(self, input_data: Dict, result: Dict) -> List[tuple]:
        """Bereitet Ergebnisdaten für Excel vor"""
        lstlzz = float(result.get("LSTLZZ", 0)) / 100
        solzlzz = float(result.get("SOLZLZZ", 0)) / 100
        bk = float(result.get("BK", 0)) / 100
        brutto = float(input_data.get("RE4", 0)) / 100

        total_tax = lstlzz + solzlzz + bk
        netto = brutto - total_tax

        return [
            ("Lohnsteuer", f"{lstlzz:.2f} €"),
            ("Solidaritätszuschlag", f"{solzlzz:.2f} €"),
            ("Kirchensteuer", f"{bk:.2f} €"),
            ("Gesamte Steuerbelastung", f"{total_tax:.2f} €"),
            ("Nettolohn", f"{netto:.2f} €"),
        ]


class ComparisonExportService:
    """Service für Vergleichsberichte"""

    def generate_comparison_report(
        self, calculations: List[Dict], format_type: str = "pdf"
    ) -> bytes:
        """Generiert einen Vergleichsbericht für mehrere Berechnungen"""
        if format_type.lower() == "pdf":
            return self._generate_pdf_comparison(calculations)
        else:
            return self._generate_excel_comparison(calculations)

    def _generate_pdf_comparison(self, calculations: List[Dict]) -> bytes:
        """Generiert PDF-Vergleichsbericht"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=18,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            fontSize=18,
            spaceAfter=30,
            textColor=colors.darkblue,
        )

        story = []

        # Titel
        title = Paragraph("Lohnsteuervergleich 2025", title_style)
        story.append(title)
        story.append(Spacer(1, 12))

        # Vergleichstabelle erstellen
        table_data = [
            [
                "Szenario",
                "Bruttolohn",
                "Lohnsteuer",
                "Solidaritätszuschlag",
                "Kirchensteuer",
                "Nettolohn",
            ]
        ]

        for i, calc in enumerate(calculations, 1):
            input_data = calc["input"]
            result = calc["result"]
            name = calc.get("name", f"Szenario {i}")

            brutto = float(input_data.get("RE4", 0)) / 100
            lstlzz = float(result.get("LSTLZZ", 0)) / 100
            solzlzz = float(result.get("SOLZLZZ", 0)) / 100
            bk = float(result.get("BK", 0)) / 100
            netto = brutto - lstlzz - solzlzz - bk

            table_data.append(
                [
                    name,
                    f"{brutto:.2f} €",
                    f"{lstlzz:.2f} €",
                    f"{solzlzz:.2f} €",
                    f"{bk:.2f} €",
                    f"{netto:.2f} €",
                ]
            )

        table = Table(
            table_data,
            colWidths=[
                1.5 * inch,
                1.2 * inch,
                1.2 * inch,
                1.2 * inch,
                1.2 * inch,
                1.2 * inch,
            ],
        )
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 10),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.lightblue),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )

        story.append(table)

        doc.build(story)
        pdf_data = buffer.getvalue()
        buffer.close()

        return pdf_data

    def _generate_excel_comparison(self, calculations: List[Dict]) -> bytes:
        """Generiert Excel-Vergleichsbericht"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Lohnsteuervergleich"

        # Header
        headers = [
            "Szenario",
            "Bruttolohn",
            "Lohnsteuer",
            "Solidaritätszuschlag",
            "Kirchensteuer",
            "Nettolohn",
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="4472C4", end_color="4472C4", fill_type="solid"
            )

        # Daten
        for row, calc in enumerate(calculations, 2):
            input_data = calc["input"]
            result = calc["result"]
            name = calc.get("name", f"Szenario {row - 1}")

            brutto = float(input_data.get("RE4", 0)) / 100
            lstlzz = float(result.get("LSTLZZ", 0)) / 100
            solzlzz = float(result.get("SOLZLZZ", 0)) / 100
            bk = float(result.get("BK", 0)) / 100
            netto = brutto - lstlzz - solzlzz - bk

            ws.cell(row=row, column=1, value=name)
            ws.cell(row=row, column=2, value=f"{brutto:.2f} €")
            ws.cell(row=row, column=3, value=f"{lstlzz:.2f} €")
            ws.cell(row=row, column=4, value=f"{solzlzz:.2f} €")
            ws.cell(row=row, column=5, value=f"{bk:.2f} €")
            ws.cell(row=row, column=6, value=f"{netto:.2f} €")

        # Spaltenbreite anpassen
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 18

        buffer = io.BytesIO()
        wb.save(buffer)
        excel_data = buffer.getvalue()
        buffer.close()

        return excel_data
